"""
Testes das regras de negócio (sem abrir a interface).

Rodar:
    python -m unittest discover testes
"""

import os
import sys
import tempfile
import unittest
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))


class TesteBase(unittest.TestCase):
    """Cada teste roda em um banco novo e temporário."""

    def setUp(self):
        self.temporario = tempfile.TemporaryDirectory()
        os.environ["CAIXA_DB"] = str(Path(self.temporario.name) / "teste.db")

        from core import db, repositorio
        db.inicializar()

        self.db = db
        self.repositorio = repositorio
        self.funcionario_id = repositorio.criar_funcionario("Ana", "Atendente")
        self.produto = repositorio.buscar_produto(
            repositorio.criar_produto("X-Burger Teste", "Lanches", 1800)
        )

    def tearDown(self):
        os.environ.pop("CAIXA_DB", None)
        self.temporario.cleanup()

    def item(self, quantidade=1, preco=None):
        return {
            "produto_id": self.produto["id"],
            "nome": self.produto["nome"],
            "categoria": self.produto["categoria"],
            "quantidade": quantidade,
            "preco_unit_centavos": preco if preco is not None else self.produto["preco_centavos"],
        }


class TesteAuth(unittest.TestCase):
    def test_verifica_senha_correta(self):
        from core.auth import gerar_hash_senha, verificar_senha
        hash_senha = gerar_hash_senha("segredo123")
        self.assertTrue(verificar_senha("segredo123", hash_senha))

    def test_recusa_senha_incorreta(self):
        from core.auth import gerar_hash_senha, verificar_senha
        hash_senha = gerar_hash_senha("segredo123")
        self.assertFalse(verificar_senha("outra-senha", hash_senha))

    def test_hash_nao_repete_para_mesma_senha(self):
        """Salt aleatório: duas senhas iguais geram hashes diferentes."""
        from core.auth import gerar_hash_senha
        self.assertNotEqual(gerar_hash_senha("segredo123"), gerar_hash_senha("segredo123"))


class TesteAcessoAdministrativo(TesteBase):
    def test_login_padrao_do_gerente_semeado_funciona(self):
        """db.inicializar() semeia o Gerente inicial com login/senha padrão."""
        from core import db, servicos
        gerente = servicos.autenticar_gerente(db.LOGIN_GERENTE_PADRAO, db.SENHA_GERENTE_PADRAO)
        self.assertEqual(gerente["cargo"], "Gerente")

    def test_autenticacao_recusa_senha_errada(self):
        from core import db, servicos
        with self.assertRaises(servicos.ErroDeNegocio):
            servicos.autenticar_gerente(db.LOGIN_GERENTE_PADRAO, "senha-errada")

    def test_criar_gerente_exige_login_e_senha(self):
        from core.servicos import ErroDeNegocio, criar_funcionario
        with self.assertRaises(ErroDeNegocio):
            criar_funcionario("Novo Gerente", "Gerente")  # sem login nem senha

    def test_criar_funcionario_nao_gerente_ignora_login_e_senha(self):
        """Login/senha só valem para o cargo Gerente — em outro cargo, são
        ignorados silenciosamente para não abrir uma porta lateral de acesso."""
        from core import repositorio, servicos

        funcionario_id = servicos.criar_funcionario(
            "Atendente Qualquer", "Atendente", login="tentativa", senha="123456"
        )
        funcionario = repositorio.buscar_funcionario(funcionario_id)
        self.assertIsNone(funcionario["login"])
        self.assertIsNone(funcionario["senha_hash"])

    def test_criar_e_autenticar_novo_gerente(self):
        from core import servicos

        servicos.criar_funcionario("Segunda Gerente", "Gerente",
                                   login="segunda.gerente", senha="troque-depois")
        gerente = servicos.autenticar_gerente("segunda.gerente", "troque-depois")
        self.assertEqual(gerente["nome"], "Segunda Gerente")

    def test_atualizar_gerente_sem_nova_senha_mantem_a_atual(self):
        from core import servicos

        funcionario_id = servicos.criar_funcionario(
            "Gerente Editado", "Gerente", login="editado", senha="senha-original"
        )
        servicos.atualizar_funcionario(funcionario_id, "Gerente Editado", "Gerente",
                                       login="editado", senha="")

        gerente = servicos.autenticar_gerente("editado", "senha-original")
        self.assertEqual(gerente["id"], funcionario_id)

    def test_editar_para_gerente_sem_senha_e_recusado(self):
        from core.servicos import ErroDeNegocio, atualizar_funcionario

        with self.assertRaises(ErroDeNegocio):
            atualizar_funcionario(self.funcionario_id, "Ana", "Gerente", login="ana.gerente")


class TesteMoeda(unittest.TestCase):
    def test_converte_formatos_digitados_pelo_usuario(self):
        from core.moeda import para_centavos
        self.assertEqual(para_centavos("12,50"), 1250)
        self.assertEqual(para_centavos("12.50"), 1250)
        self.assertEqual(para_centavos("R$ 1.234,56"), 123456)
        self.assertEqual(para_centavos(7), 700)
        self.assertEqual(para_centavos(0), 0)

    def test_recusa_texto_invalido(self):
        from core.moeda import ValorInvalido, para_centavos
        for entrada in ("abc", "", "12,,5"):
            with self.assertRaises(ValorInvalido):
                para_centavos(entrada)

    def test_formata_em_padrao_brasileiro(self):
        from core.moeda import formatar
        self.assertEqual(formatar(1250), "R$ 12,50")
        self.assertEqual(formatar(123456), "R$ 1.234,56")
        self.assertEqual(formatar(0), "R$ 0,00")

    def test_soma_em_centavos_nao_perde_precisao(self):
        """O clássico 0.1 + 0.2 != 0.3 que quebra fechamento de caixa."""
        from core.moeda import para_centavos
        total = sum(para_centavos("0,10") for _ in range(10))
        self.assertEqual(total, para_centavos("1,00"))


class TesteAberturaFechamento(TesteBase):
    def test_nao_permite_dois_caixas_abertos(self):
        from core.servicos import ErroDeNegocio, abrir_caixa
        abrir_caixa(self.funcionario_id, 10000)
        with self.assertRaises(ErroDeNegocio):
            abrir_caixa(self.funcionario_id, 5000)

    def test_venda_sem_caixa_aberto_e_recusada(self):
        from core.servicos import ErroDeNegocio, registrar_venda
        with self.assertRaises(ErroDeNegocio):
            registrar_venda(self.funcionario_id, [self.item()])

    def test_fechamento_calcula_diferenca_da_gaveta(self):
        from core.servicos import abrir_caixa, fechar_caixa, registrar_venda

        caixa_id = abrir_caixa(self.funcionario_id, 10000)          # R$ 100,00 de troco
        registrar_venda(self.funcionario_id, [self.item(2)], "Dinheiro")   # R$ 36,00
        registrar_venda(self.funcionario_id, [self.item(1)], "PIX")        # R$ 18,00

        resumo, _ = fechar_caixa(caixa_id, self.funcionario_id,
                                 valor_contado_centavos=13600, gerar_planilha=False)

        self.assertEqual(resumo["total_centavos"], 5400)             # 36 + 18
        self.assertEqual(resumo["total_dinheiro_centavos"], 3600)    # só o dinheiro
        self.assertEqual(resumo["esperado_gaveta_centavos"], 13600)  # 100 + 36
        self.assertEqual(resumo["caixa"]["diferenca_centavos"], 0)
        self.assertEqual(resumo["caixa"]["status"], "FECHADO")

    def test_fechamento_registra_quebra_de_caixa(self):
        from core.servicos import abrir_caixa, fechar_caixa, registrar_venda

        caixa_id = abrir_caixa(self.funcionario_id, 0)
        registrar_venda(self.funcionario_id, [self.item(1)], "Dinheiro")  # R$ 18,00
        resumo, _ = fechar_caixa(caixa_id, self.funcionario_id,
                                 valor_contado_centavos=1700, gerar_planilha=False)

        self.assertEqual(resumo["caixa"]["diferenca_centavos"], -100)  # faltou R$ 1,00

    def test_nao_fecha_duas_vezes(self):
        from core.servicos import ErroDeNegocio, abrir_caixa, fechar_caixa
        caixa_id = abrir_caixa(self.funcionario_id, 0)
        fechar_caixa(caixa_id, self.funcionario_id, 0, gerar_planilha=False)
        with self.assertRaises(ErroDeNegocio):
            fechar_caixa(caixa_id, self.funcionario_id, 0, gerar_planilha=False)


class TesteVendas(TesteBase):
    def test_registra_venda_com_itens_e_horario(self):
        from core.servicos import abrir_caixa, registrar_venda

        caixa_id = abrir_caixa(self.funcionario_id, 0)
        venda = registrar_venda(self.funcionario_id, [self.item(3)], "Cartão de Débito")

        self.assertEqual(venda["total_centavos"], 5400)
        itens = self.repositorio.itens_da_venda(venda["venda_id"])
        self.assertEqual(len(itens), 1)
        self.assertEqual(itens[0]["quantidade"], 3)
        self.assertEqual(itens[0]["nome_produto"], "X-Burger Teste")
        self.assertRegex(venda["data_hora"], r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$")

    def test_venda_vazia_e_recusada(self):
        from core.servicos import ErroDeNegocio, abrir_caixa, registrar_venda
        abrir_caixa(self.funcionario_id, 0)
        with self.assertRaises(ErroDeNegocio):
            registrar_venda(self.funcionario_id, [])

    def test_venda_cancelada_sai_do_total(self):
        from core.servicos import (abrir_caixa, cancelar_venda, registrar_venda,
                                   resumo_caixa)

        caixa_id = abrir_caixa(self.funcionario_id, 0)
        registrar_venda(self.funcionario_id, [self.item(1)])
        venda = registrar_venda(self.funcionario_id, [self.item(1)])
        cancelar_venda(venda["venda_id"], "Cliente desistiu")

        resumo = resumo_caixa(caixa_id)
        self.assertEqual(resumo["qtd_vendas"], 1)
        self.assertEqual(resumo["total_centavos"], 1800)
        self.assertEqual(resumo["total_cancelado_centavos"], 1800)

    def test_preco_do_item_e_congelado_na_venda(self):
        """Mudar o preço no cardápio não pode alterar o histórico."""
        from core.servicos import abrir_caixa, registrar_venda, resumo_caixa

        caixa_id = abrir_caixa(self.funcionario_id, 0)
        registrar_venda(self.funcionario_id, [self.item(1)])
        self.repositorio.atualizar_produto(self.produto["id"], "X-Burger Teste", "Lanches", 9900)

        self.assertEqual(resumo_caixa(caixa_id)["total_centavos"], 1800)


class TesteVendaDividida(TesteBase):
    """Cliente pagando parte em dinheiro, parte em outra forma."""

    def _pagamentos(self, valor1=2000, valor2=1600, forma1="Dinheiro", forma2="PIX"):
        return [
            {"forma_pagamento": forma1, "valor_centavos": valor1},
            {"forma_pagamento": forma2, "valor_centavos": valor2},
        ]

    def test_forma_de_pagamento_combina_as_duas_formas(self):
        from core.servicos import abrir_caixa, registrar_venda

        caixa_id = abrir_caixa(self.funcionario_id, 0)
        venda = registrar_venda(self.funcionario_id, [self.item(2)],
                                pagamentos=self._pagamentos())  # total 3600 = 2000 + 1600

        self.assertEqual(venda["total_centavos"], 3600)
        registrada = next(v for v in self.repositorio.listar_vendas(caixa_id)
                          if v["id"] == venda["venda_id"])
        self.assertEqual(registrada["forma_pagamento"], "Dinheiro/PIX")

    def test_grava_uma_linha_por_forma_em_pagamentos_venda(self):
        from core.servicos import abrir_caixa, registrar_venda

        abrir_caixa(self.funcionario_id, 0)
        venda = registrar_venda(self.funcionario_id, [self.item(2)],
                                pagamentos=self._pagamentos())

        pagamentos = self.repositorio.pagamentos_da_venda(venda["venda_id"])
        self.assertEqual(len(pagamentos), 2)
        self.assertEqual({p["forma_pagamento"] for p in pagamentos}, {"Dinheiro", "PIX"})
        self.assertEqual(sum(p["valor_centavos"] for p in pagamentos), 3600)

    def test_soma_incorreta_e_recusada(self):
        from core.servicos import ErroDeNegocio, abrir_caixa, registrar_venda

        abrir_caixa(self.funcionario_id, 0)
        with self.assertRaises(ErroDeNegocio):
            registrar_venda(self.funcionario_id, [self.item(2)],  # total 3600
                            pagamentos=self._pagamentos(valor1=2000, valor2=1000))  # soma 3000

    def test_exige_duas_formas_diferentes(self):
        from core.servicos import ErroDeNegocio, abrir_caixa, registrar_venda

        abrir_caixa(self.funcionario_id, 0)
        with self.assertRaises(ErroDeNegocio):
            registrar_venda(self.funcionario_id, [self.item(2)],
                            pagamentos=self._pagamentos(forma1="Dinheiro", forma2="Dinheiro"))

    def test_exige_valores_positivos(self):
        from core.servicos import ErroDeNegocio, abrir_caixa, registrar_venda

        abrir_caixa(self.funcionario_id, 0)
        with self.assertRaises(ErroDeNegocio):
            registrar_venda(self.funcionario_id, [self.item(2)],
                            pagamentos=self._pagamentos(valor1=3600, valor2=0))

    def test_aceita_exatamente_duas_formas(self):
        from core.servicos import ErroDeNegocio, abrir_caixa, registrar_venda

        abrir_caixa(self.funcionario_id, 0)
        with self.assertRaises(ErroDeNegocio):
            registrar_venda(self.funcionario_id, [self.item(2)],
                            pagamentos=[{"forma_pagamento": "Dinheiro", "valor_centavos": 3600}])

    def test_gaveta_recebe_so_a_parte_em_dinheiro(self):
        """A conferência de gaveta não pode contar o valor pago em PIX."""
        from core.servicos import abrir_caixa, fechar_caixa, registrar_venda

        caixa_id = abrir_caixa(self.funcionario_id, 0)
        registrar_venda(self.funcionario_id, [self.item(2)],
                        pagamentos=self._pagamentos(valor1=2000, valor2=1600))

        resumo, _ = fechar_caixa(caixa_id, self.funcionario_id,
                                 valor_contado_centavos=2000, gerar_planilha=False)

        self.assertEqual(resumo["total_dinheiro_centavos"], 2000)
        self.assertEqual(resumo["esperado_gaveta_centavos"], 2000)
        self.assertEqual(resumo["caixa"]["diferenca_centavos"], 0)


class TestePlanilha(TesteBase):
    def test_gera_arquivo_com_as_quatro_abas(self):
        from openpyxl import load_workbook

        from core.servicos import abrir_caixa, fechar_caixa, registrar_venda

        caixa_id = abrir_caixa(self.funcionario_id, 5000)
        registrar_venda(self.funcionario_id, [self.item(2)], "Dinheiro")

        destino = Path(self.temporario.name)
        from core import planilha
        _, caminho = fechar_caixa(caixa_id, self.funcionario_id, 8600, gerar_planilha=False)
        caminho = planilha.gerar(caixa_id, pasta_destino=destino)

        self.assertTrue(caminho.exists())
        livro = load_workbook(caminho)
        self.assertEqual(livro.sheetnames,
                         ["Resumo", "Vendas do Dia", "Por Produto", "Por Funcionário"])
        self.assertEqual(livro["Vendas do Dia"]["D4"].value, "X-Burger Teste")
        self.assertEqual(livro["Vendas do Dia"]["F4"].value, 2)


class TesteHistorico(TesteBase):
    def test_caixa_fechado_aparece_na_consulta(self):
        from core.servicos import abrir_caixa, fechar_caixa, registrar_venda

        caixa_id = abrir_caixa(self.funcionario_id, 0)
        registrar_venda(self.funcionario_id, [self.item(2)])
        fechar_caixa(caixa_id, self.funcionario_id, 3600, gerar_planilha=False)

        historico = self.repositorio.listar_caixas()
        self.assertEqual(len(historico), 1)
        self.assertEqual(historico[0]["total_centavos"], 3600)
        self.assertEqual(historico[0]["qtd_vendas"], 1)


if __name__ == "__main__":
    unittest.main(verbosity=2)
