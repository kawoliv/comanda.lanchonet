"""
Geração da planilha (.xlsx) do fechamento de caixa.

São quatro abas:
    Resumo         - cabeçalho do dia, totais e conferência da gaveta
    Vendas do Dia  - uma linha por produto vendido (hora, vendedor, qtd, valor)
    Por Produto    - consolidado: quanto saiu de cada item
    Por Funcionário- quanto cada atendente vendeu

Os totais são gravados como FÓRMULAS (=SUM/=SUBTOTAL) e não como números
prontos: assim o gerente pode filtrar ou corrigir uma linha e a planilha
recalcula sozinha.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core import servicos
from core.db import PASTA_RELATORIOS
from core.moeda import para_reais

FONTE = "Arial"
MOEDA = 'R$ #,##0.00'

COR_TITULO = "12263A"   # azul escuro
COR_CABECALHO = "E8590C"  # laranja
COR_FAIXA = "F4F6F8"

_BORDA_FINA = Side(style="thin", color="C9D2DC")
BORDA = Border(left=_BORDA_FINA, right=_BORDA_FINA, top=_BORDA_FINA, bottom=_BORDA_FINA)


# --------------------------------------------------------------------------- #
# Auxiliares de formatação
# --------------------------------------------------------------------------- #


def _titulo(aba, texto: str, colunas: int, linha: int = 1) -> None:
    aba.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=colunas)
    celula = aba.cell(row=linha, column=1, value=texto)
    celula.font = Font(name=FONTE, size=14, bold=True, color="FFFFFF")
    celula.fill = PatternFill("solid", fgColor=COR_TITULO)
    celula.alignment = Alignment(horizontal="center", vertical="center")
    aba.row_dimensions[linha].height = 26


def _cabecalho(aba, linha: int, colunas: list) -> None:
    for indice, nome in enumerate(colunas, start=1):
        celula = aba.cell(row=linha, column=indice, value=nome)
        celula.font = Font(name=FONTE, size=11, bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = BORDA
    aba.row_dimensions[linha].height = 22


def _larguras(aba, larguras: list) -> None:
    for indice, largura in enumerate(larguras, start=1):
        aba.column_dimensions[get_column_letter(indice)].width = largura


def _linha_rotulo(aba, linha: int, rotulo: str, valor, moeda: bool = False, negrito: bool = False):
    celula_rotulo = aba.cell(row=linha, column=1, value=rotulo)
    celula_rotulo.font = Font(name=FONTE, size=11, bold=True)

    celula_valor = aba.cell(row=linha, column=2, value=valor)
    celula_valor.font = Font(name=FONTE, size=11, bold=negrito)
    if moeda:
        celula_valor.number_format = MOEDA
    return celula_valor


def _estilo_corpo(aba, primeira_linha: int, ultima_linha: int, colunas: int,
                  colunas_moeda=(), colunas_centro=()):
    for linha in range(primeira_linha, ultima_linha + 1):
        faixa = (linha - primeira_linha) % 2 == 1
        for coluna in range(1, colunas + 1):
            celula = aba.cell(row=linha, column=coluna)
            celula.font = Font(name=FONTE, size=10)
            celula.border = BORDA
            if faixa:
                celula.fill = PatternFill("solid", fgColor=COR_FAIXA)
            if coluna in colunas_moeda:
                celula.number_format = MOEDA
                celula.alignment = Alignment(horizontal="right")
            elif coluna in colunas_centro:
                celula.alignment = Alignment(horizontal="center")


def _linha_total(aba, linha: int, colunas: int, rotulo_coluna: int, formulas: dict) -> None:
    """`formulas` mapeia número da coluna -> string da fórmula."""
    for coluna in range(1, colunas + 1):
        celula = aba.cell(row=linha, column=coluna)
        celula.fill = PatternFill("solid", fgColor="DDE3EA")
        celula.font = Font(name=FONTE, size=11, bold=True)
        celula.border = BORDA

    aba.cell(row=linha, column=rotulo_coluna, value="TOTAL")
    for coluna, formula in formulas.items():
        celula = aba.cell(row=linha, column=coluna, value=formula)
        celula.font = Font(name=FONTE, size=11, bold=True)
        if coluna != rotulo_coluna:
            celula.number_format = MOEDA if "subtotal" not in formula.lower() else MOEDA


# --------------------------------------------------------------------------- #
# Abas
# --------------------------------------------------------------------------- #


def _aba_resumo(aba, resumo: dict) -> None:
    caixa = resumo["caixa"]
    aba.sheet_view.showGridLines = False
    _larguras(aba, [34, 24, 16, 16, 16])
    _titulo(aba, "FECHAMENTO DE CAIXA", 5)

    linha = 3
    _linha_rotulo(aba, linha, "Caixa nº", caixa["id"]); linha += 1
    _linha_rotulo(aba, linha, "Data de operação", _data_br(caixa["data_operacao"])); linha += 1
    _linha_rotulo(aba, linha, "Aberto em", caixa["aberto_em"]); linha += 1
    _linha_rotulo(aba, linha, "Aberto por", caixa["operador_abertura"]); linha += 1
    _linha_rotulo(aba, linha, "Fechado em", caixa["fechado_em"] or "-"); linha += 1
    _linha_rotulo(aba, linha, "Fechado por", caixa["operador_fechamento"] or "-"); linha += 2

    aba.cell(row=linha, column=1, value="MOVIMENTO DO DIA").font = Font(
        name=FONTE, size=12, bold=True, color=COR_TITULO
    )
    linha += 1

    _linha_rotulo(aba, linha, "Vendas registradas", resumo["qtd_vendas"]); linha += 1
    _linha_rotulo(aba, linha, "Itens vendidos", resumo["qtd_itens"]); linha += 1
    _linha_rotulo(aba, linha, "Total vendido",
                  para_reais(resumo["total_centavos"]), moeda=True, negrito=True); linha += 1
    _linha_rotulo(aba, linha, "Ticket médio",
                  para_reais(resumo["ticket_medio_centavos"]), moeda=True); linha += 1
    _linha_rotulo(aba, linha, "Vendas canceladas", resumo["qtd_canceladas"]); linha += 1
    _linha_rotulo(aba, linha, "Valor cancelado (não entra no total)",
                  para_reais(resumo["total_cancelado_centavos"]), moeda=True); linha += 2

    aba.cell(row=linha, column=1, value="POR FORMA DE PAGAMENTO").font = Font(
        name=FONTE, size=12, bold=True, color=COR_TITULO
    )
    linha += 1

    inicio_pagamento = linha + 1
    _cabecalho(aba, linha, ["Forma de pagamento", "Qtd. vendas", "Total"])
    linha += 1
    for pagamento in resumo["por_pagamento"]:
        aba.cell(row=linha, column=1, value=pagamento["forma_pagamento"])
        aba.cell(row=linha, column=2, value=int(pagamento["qtd_vendas"]))
        aba.cell(row=linha, column=3, value=para_reais(pagamento["total_centavos"]))
        linha += 1

    if resumo["por_pagamento"]:
        fim_pagamento = linha - 1
        _estilo_corpo(aba, inicio_pagamento, fim_pagamento, 3,
                      colunas_moeda=(3,), colunas_centro=(2,))
        _linha_total(aba, linha, 3, 1, {
            2: f"=SUM(B{inicio_pagamento}:B{fim_pagamento})",
            3: f"=SUM(C{inicio_pagamento}:C{fim_pagamento})",
        })
        aba.cell(row=linha, column=2).number_format = "0"
        linha += 1
    linha += 1

    aba.cell(row=linha, column=1, value="CONFERÊNCIA DA GAVETA").font = Font(
        name=FONTE, size=12, bold=True, color=COR_TITULO
    )
    linha += 1

    linha_abertura = linha
    _linha_rotulo(aba, linha, "(+) Troco inicial (abertura)",
                  para_reais(caixa["valor_abertura_centavos"]), moeda=True); linha += 1
    linha_dinheiro = linha
    _linha_rotulo(aba, linha, "(+) Recebido em dinheiro",
                  para_reais(resumo["total_dinheiro_centavos"]), moeda=True); linha += 1

    linha_esperado = linha
    celula = _linha_rotulo(aba, linha, "(=) Esperado na gaveta",
                           f"=B{linha_abertura}+B{linha_dinheiro}", moeda=True, negrito=True)
    celula.number_format = MOEDA
    linha += 1

    linha_contado = linha
    _linha_rotulo(aba, linha, "(-) Contado no fechamento",
                  para_reais(caixa["valor_contado_centavos"] or 0), moeda=True); linha += 1

    celula = _linha_rotulo(aba, linha, "(=) Diferença",
                           f"=B{linha_contado}-B{linha_esperado}", moeda=True, negrito=True)
    celula.number_format = MOEDA
    linha += 2

    if caixa["observacoes"]:
        aba.cell(row=linha, column=1, value="Observações").font = Font(name=FONTE, bold=True)
        celula = aba.cell(row=linha, column=2, value=caixa["observacoes"])
        celula.alignment = Alignment(wrap_text=True, vertical="top")
        linha += 2

    aba.cell(
        row=linha, column=1,
        value=f"Gerado automaticamente pelo Sistema de Caixa em "
              f"{datetime.now().strftime('%d/%m/%Y %H:%M')}.",
    ).font = Font(name=FONTE, size=9, italic=True, color="6B7785")


def _aba_vendas(aba, resumo: dict) -> None:
    colunas = ["Venda nº", "Hora", "Vendedor", "Produto", "Categoria",
               "Qtd.", "Valor unit.", "Subtotal", "Pagamento"]
    _larguras(aba, [10, 10, 20, 28, 16, 8, 14, 14, 18])
    _titulo(aba, "VENDAS DO DIA — DETALHADO POR PRODUTO", len(colunas))
    _cabecalho(aba, 3, colunas)

    linha = 4
    for item in resumo["itens"]:
        aba.cell(row=linha, column=1, value=int(item["venda_id"]))
        aba.cell(row=linha, column=2, value=item["data_hora"][11:16])
        aba.cell(row=linha, column=3, value=item["vendedor"])
        aba.cell(row=linha, column=4, value=item["nome_produto"])
        aba.cell(row=linha, column=5, value=item["categoria"])
        aba.cell(row=linha, column=6, value=int(item["quantidade"]))
        aba.cell(row=linha, column=7, value=para_reais(item["preco_unit_centavos"]))
        aba.cell(row=linha, column=8, value=para_reais(item["subtotal_centavos"]))
        aba.cell(row=linha, column=9, value=item["forma_pagamento"])
        linha += 1

    if resumo["itens"]:
        _estilo_corpo(aba, 4, linha - 1, len(colunas),
                      colunas_moeda=(7, 8), colunas_centro=(1, 2, 6))
        _linha_total(aba, linha, len(colunas), 5, {
            6: f"=SUM(F4:F{linha - 1})",
            8: f"=SUM(H4:H{linha - 1})",
        })
        aba.cell(row=linha, column=6).number_format = "0"
        aba.auto_filter.ref = f"A3:I{linha - 1}"

    aba.freeze_panes = "A4"


def _aba_por_produto(aba, resumo: dict) -> None:
    colunas = ["Produto", "Categoria", "Qtd. vendida", "Valor unit. médio", "Total vendido"]
    _larguras(aba, [30, 18, 14, 18, 16])
    _titulo(aba, "CONSOLIDADO POR PRODUTO", len(colunas))
    _cabecalho(aba, 3, colunas)

    linha = 4
    for produto in resumo["por_produto"]:
        quantidade = int(produto["quantidade"])
        total = int(produto["total_centavos"])
        aba.cell(row=linha, column=1, value=produto["nome_produto"])
        aba.cell(row=linha, column=2, value=produto["categoria"])
        aba.cell(row=linha, column=3, value=quantidade)
        aba.cell(row=linha, column=4, value=para_reais(total / quantidade) if quantidade else 0)
        aba.cell(row=linha, column=5, value=para_reais(total))
        linha += 1

    if resumo["por_produto"]:
        _estilo_corpo(aba, 4, linha - 1, len(colunas),
                      colunas_moeda=(4, 5), colunas_centro=(3,))
        _linha_total(aba, linha, len(colunas), 1, {
            3: f"=SUM(C4:C{linha - 1})",
            5: f"=SUM(E4:E{linha - 1})",
        })
        aba.cell(row=linha, column=3).number_format = "0"

    aba.freeze_panes = "A4"


def _aba_por_funcionario(aba, resumo: dict) -> None:
    colunas = ["Vendedor", "Qtd. vendas", "Total vendido", "Ticket médio", "% do dia"]
    _larguras(aba, [26, 14, 16, 16, 12])
    _titulo(aba, "VENDAS POR FUNCIONÁRIO", len(colunas))
    _cabecalho(aba, 3, colunas)

    linha = 4
    for funcionario in resumo["por_funcionario"]:
        quantidade = int(funcionario["qtd_vendas"])
        total = int(funcionario["total_centavos"])
        aba.cell(row=linha, column=1, value=funcionario["vendedor"])
        aba.cell(row=linha, column=2, value=quantidade)
        aba.cell(row=linha, column=3, value=para_reais(total))
        aba.cell(row=linha, column=4, value=para_reais(total / quantidade) if quantidade else 0)
        linha += 1

    ultima = linha - 1
    if resumo["por_funcionario"]:
        for atual in range(4, linha):
            celula = aba.cell(row=atual, column=5, value=f"=IFERROR(C{atual}/$C${linha},0)")
            celula.number_format = "0.0%"

        _estilo_corpo(aba, 4, ultima, len(colunas),
                      colunas_moeda=(3, 4), colunas_centro=(2, 5))
        for atual in range(4, linha):
            aba.cell(row=atual, column=5).number_format = "0.0%"

        _linha_total(aba, linha, len(colunas), 1, {
            2: f"=SUM(B4:B{ultima})",
            3: f"=SUM(C4:C{ultima})",
        })
        aba.cell(row=linha, column=2).number_format = "0"

    aba.freeze_panes = "A4"


# --------------------------------------------------------------------------- #
# Entrada pública
# --------------------------------------------------------------------------- #


def _data_br(data_iso: str) -> str:
    try:
        return datetime.strptime(data_iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return data_iso or "-"


def gerar(caixa_id: int, pasta_destino: Path = None) -> Path:
    """Monta o arquivo .xlsx do caixa informado e devolve o caminho salvo."""
    resumo = servicos.resumo_caixa(caixa_id)
    caixa = resumo["caixa"]

    pasta = Path(pasta_destino) if pasta_destino else PASTA_RELATORIOS
    pasta.mkdir(parents=True, exist_ok=True)

    arquivo = pasta / f"fechamento_{caixa['data_operacao']}_caixa-{caixa['id']}.xlsx"

    livro = Workbook()
    _aba_resumo(livro.active, resumo)
    livro.active.title = "Resumo"
    _aba_vendas(livro.create_sheet("Vendas do Dia"), resumo)
    _aba_por_produto(livro.create_sheet("Por Produto"), resumo)
    _aba_por_funcionario(livro.create_sheet("Por Funcionário"), resumo)

    livro.save(arquivo)
    return arquivo
